import base64
import io
import tempfile
from datetime import date, timedelta
from flask import Flask, render_template, request, send_file
from openpyxl import Workbook
from openpyxl.styles import PatternFill
from openpyxl.drawing.image import Image as XLImage
from logic.data import taches, parser_taches
from logic.calcul import passe_avant, passe_arriere, calculer_marges, liste_chemin_critique, valider_taches, generer_conseils

app = Flask(__name__)


def preparer_graphe(taches, resultats):
    nodes = []
    for nom, infos in taches.items():
        couleur = "#ff6b6b" if resultats[nom]["critique"] else "#97c2fc"
        nodes.append({
            "id": nom,
            "label": f"{nom}\n({infos['duree']}j)",
            "color": couleur
        })
    edges = []
    for nom, infos in taches.items():
        for dep in infos["depend_de"]:
            edges.append({"from": dep, "to": nom, "arrows": "to"})
    return nodes, edges


def preparer_gantt(taches, resultats, date_debut=None):
    if date_debut is None:
        date_debut = date.today()
    taches_gantt = []
    for nom, infos in taches.items():
        r = resultats[nom]
        debut = date_debut + timedelta(days=r["ES"])
        fin = date_debut + timedelta(days=r["EF"])
        taches_gantt.append({
            "id": nom,
            "name": f"{nom} - {infos['description']}",
            "start": debut.isoformat(),
            "end": fin.isoformat(),
            "progress": 0,
            "dependencies": ",".join(infos["depend_de"]),
            "custom_class": "critique" if r["critique"] else ""
        })
    return taches_gantt


def decoder_et_sauver_image(image_b64):
    en_tete, donnees = image_b64.split(',', 1)
    donnees_binaires = base64.b64decode(donnees)
    fichier_temp = tempfile.NamedTemporaryFile(delete=False, suffix=".png")
    fichier_temp.write(donnees_binaires)
    fichier_temp.close()
    return fichier_temp.name


@app.route('/')
def accueil():
    return render_template('accueil.html')


@app.route('/demo')
def demo():
    av = passe_avant(taches)
    ar = passe_arriere(taches, av)
    resultats = calculer_marges(taches, av, ar)
    chemin_critique = liste_chemin_critique(resultats)
    duree_totale = max(r["EF"] for r in resultats.values())
    nodes, edges = preparer_graphe(taches, resultats)
    taches_gantt = preparer_gantt(taches, resultats)
    return render_template(
        'resultats.html',
        taches=taches,
        resultats=resultats,
        chemin_critique=chemin_critique,
        duree_totale=duree_totale,
        nodes=nodes,
        edges=edges,
        taches_gantt=taches_gantt,
        date_debut=None
    )


@app.route('/saisie')
def saisie():
    return render_template('saisie.html')


@app.route('/resultats_custom', methods=['POST'])
def resultats_custom():
    texte = request.form['taches']
    date_debut_str = request.form.get('date_debut', '').strip()
    date_fin_str = request.form.get('date_fin_cible', '').strip()

    # Conversion date_debut
    try:
        date_debut_obj = date.fromisoformat(date_debut_str) if date_debut_str else date.today()
    except ValueError:
        date_debut_obj = date.today()

    taches_custom, erreurs = parser_taches(texte)

    if not taches_custom and not erreurs:
        erreurs.append("Aucune tâche saisie.")
    if not erreurs:
        erreurs += valider_taches(taches_custom)
    if erreurs:
        return render_template('erreur.html', erreurs=erreurs)

    av = passe_avant(taches_custom)
    ar = passe_arriere(taches_custom, av)
    resultats = calculer_marges(taches_custom, av, ar)
    chemin_critique = liste_chemin_critique(resultats)
    duree_totale = max(r["EF"] for r in resultats.values())
    nodes, edges = preparer_graphe(taches_custom, resultats)
    taches_gantt = preparer_gantt(taches_custom, resultats, date_debut_obj)

    # Date de fin affichage JJ/MM/AAAA + objet date pour le calcul des conseils
    date_fin_affichage = None
    date_fin_obj = None
    if date_fin_str:
        try:
            date_fin_obj = date.fromisoformat(date_fin_str)
            date_fin_affichage = date_fin_obj.strftime('%d/%m/%Y')
        except ValueError:
            pass

    conseils = generer_conseils(
        taches_custom, resultats, chemin_critique, duree_totale,
        date_debut=date_debut_obj, date_fin_cible=date_fin_obj
    )

    return render_template(
        'resultats.html',
        taches=taches_custom,
        resultats=resultats,
        chemin_critique=chemin_critique,
        duree_totale=duree_totale,
        nodes=nodes,
        edges=edges,
        taches_gantt=taches_gantt,
        date_debut=date_debut_str,
        date_fin=date_fin_affichage,
        conseils=conseils
    )


@app.route('/export_excel', methods=['POST'])
def export_excel():
    data = request.get_json()
    taches_data = data['taches']
    resultats_data = data['resultats']
    duree_totale = data['duree_totale']
    chemin_critique = data['chemin_critique']

    wb = Workbook()
    ws = wb.active
    ws.title = "Résultats"
    ws.append(["Tâche", "Description", "Durée", "ES", "EF", "LS", "LF", "Marge", "Critique"])
    rouge = PatternFill(start_color="FFCCCC", end_color="FFCCCC", fill_type="solid")

    for nom, r in resultats_data.items():
        infos = taches_data[nom]
        ws.append([nom, infos["description"], infos["duree"],
                   r["ES"], r["EF"], r["LS"], r["LF"], r["marge"],
                   "OUI" if r["critique"] else "Non"])
        if r["critique"]:
            for cell in ws[ws.max_row]:
                cell.fill = rouge

    ws.append([])
    ws.append([f"Durée totale du projet : {duree_totale} jours"])
    ws.append([f"Chemin critique : {' -> '.join(chemin_critique)}"])

    ws_graphe = wb.create_sheet("Graphe PERT")
    ws_graphe.add_image(XLImage(decoder_et_sauver_image(data['image_graphe'])), "A1")

    ws_gantt = wb.create_sheet("Diagramme Gantt")
    ws_gantt.add_image(XLImage(decoder_et_sauver_image(data['image_gantt'])), "A1")

    sortie = io.BytesIO()
    wb.save(sortie)
    sortie.seek(0)
    return send_file(
        sortie,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        as_attachment=True,
        download_name="resultats_pert_cpm.xlsx"
    )


if __name__ == '__main__':
    app.run(debug=True)